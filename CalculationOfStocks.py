import numpy as np
import pandas as pd
from openpyxl import Workbook,load_workbook
import time

data=["APPLE","ALLIANZ","GENERAL ELECTRIC","LLOYDS BANKING GROUP","BT GROUP","DEUTSCHE POST","JOHNSON & JOHNSON"]
index = [0, 1, 2, 3, 4, 5, 6]
my_series = pd.Series(data,index)
pandas_series = pd.Series(data=index, index=data, dtype=int)
print("Calculate an index value of the following stocks:(USD)")
print(pandas_series)
wf = load_workbook("FX_EUR_GBP.xlsx")
wfs = wf.active
wd = load_workbook("StaticData_Apr20.xlsx")
wds = wd.active
wt = load_workbook("TimeSeriesData_Apr15-Apr20.xlsx")
wts = wt.active

def calcMarketCapWeighted(t,n,p,s,f,c,x,M,D):
    add= 0
    for a in n:
        add += p * s * f * c * x
    Mt = add
    result = Mt / D
    return result

def calcPriceWeightedWeightingFac(t,n,p,w,c,x,M,D):
    add = 0
    for a in n:
        add += p * w * c * x
    Mt = add
    result = Mt / D
    return result


def calcIndexDivisor(Dt,D,n,p,s,f,c,x,DeltaMC):
    add = 0
    for a in n:
        add += (p * s * f * c * x)+DeltaMC
    Mt = add
    for b in n:
        add += p * s * f * c * x
    Mt1 = add

    result = (Mt / Mt1)* D
    return result


def euroToUsd():
    pass


def gbpToUsd():
    pass


def FX_EURO_GBP():
    list2 = []
    index = 0
    for row in range(2, wfs.max_row + 1):
        for cell in range(1, wfs.max_column + 1):
            list2.insert(index, str(wfs.cell(row,cell).value))
            index += 1
    return list2
def StaticData_Apr20():
    list1 = []
    index = 0
    for row in wds.iter_rows(min_row=2, min_col=1, max_row=8, max_col=4):
        for cell in row:
            list1.insert(index,cell.value)
            index +=1
    return list1


def TimeSeriesData_Apr15():
    list3 = []
    index = 0
    for row in range(1,1):
        for cell in range(1, wts.max_column + 1):
            list3.insert(index, cell.value)
            index += 1
            #print(" | " + str(wts.cell(satir, sutun).value) + " | ", end="")
        #print()
    return list3

while True:
    print("Enter a Position:\n  .Analsyt  \n  .System Administrator")
    position = input("Enter a Positon:")
    if(position == "Analyst"):
        print("Would you like to index value calcuated on daily basisi?")
        selection = input("Y/N")
        if(selection=="Y"):
            pass
        elif(selection=="N"):
            pass
        else:
            print("Invalid Operation")
            continue
        print("Would you like to index value chart on daily basisi?")
        selection = input("Y/N")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            pass
        else:
            print("Invalid Operation")
            continue
    elif(position == "System Administrator"):
        print("Would you like to a script to deploy the entire application?")
        selection = input("Y/N")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            pass
        else:
            print("Invalid Operation")
            continue
    else:
        print("Invalid Operation")


    print("Operations: \n 1.Select An Index For Calculation: \n 2.Exit Press Q")
    operations = int(input("Enter: "))
    if(operations == 0):
        pandas_series[operations]
        print("Selected Index: {} , Product : {}".format(operations,pandas_series.index[operations]))
        print("Static Datas Information:")
        time.sleep(2)
        a = StaticData_Apr20()
        seperated = a[0:4]
        print(seperated)
        if (seperated[3] == "USD"):
            pass
        elif (seperated[3] == "EUR"):
            pass
        elif (seperated[3] == "GBP"):
            pass
        else:
            print("Invalid Opreration")
            continue

        print("FX_Euro_Gbp Datas Information:")
        time.sleep(2)
        b = FX_EURO_GBP()
        print(b)
        print("Time Series Datas Information:")
        TimeSeriesData_Apr15()
        time.sleep(2)
        print("Do you want to Calculate?")
        selection = input("Y/N :")
        if(selection=="Y"):
            pass
        elif(selection=="N"):
            continue
        else:
            print("Invalid Selection")
    elif(operations == 1 ):
        pandas_series[operations]
        print("Selected Index: {} , Product : {}".format(operations, pandas_series.index[operations]))
        print("Static Datas Information:")
        time.sleep(2)
        a = StaticData_Apr20()
        seperated = a[4:8]
        print(seperated)
        if (seperated[3] == "USD"):
            pass
        elif (seperated[3] == "EUR"):
            pass
        elif (seperated[3] == "GBP"):
            pass
        else:
            print("Invalid Opreration")
            continue
        print("FX_Euro_Gbp Datas Information:")
        time.sleep(2)
        b = FX_EURO_GBP()
        print(b)
        print("Time Series Datas Information:")
        time.sleep(2)
        c=TimeSeriesData_Apr15()
        print("Do you want to Calculate?")
        selection = input("Y/N :")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            continue
        else:
            print("Invalid Selection")
    elif (operations == 2):
        pandas_series[operations]
        print("Selected Index: {} , Product : {}".format(operations, pandas_series.index[operations]))
        print("Static Datas Information:")
        time.sleep(2)
        a = StaticData_Apr20()
        seperated = a[8:12]
        print(seperated)
        if (seperated[3] == "USD"):
            pass
        elif (seperated[3] == "EUR"):
            pass
        elif (seperated[3] == "GBP"):
            pass
        else:
            print("Invalid Opreration")
            continue
        print("FX_Euro_Gbp Datas Information:")
        time.sleep(2)
        b = FX_EURO_GBP()
        print(b)
        print("Time Series Datas Information:")
        time.sleep(2)
        TimeSeriesData_Apr15()
        print("Do you want to Calculate?")
        selection = input("Y/N :")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            continue
        else:
            print("Invalid Selection")
    elif (operations == 3):
        pandas_series[operations]
        print("Selected Index: {} , Product : {}".format(operations, pandas_series.index[operations]))
        print("Static Datas Information:")
        time.sleep(2)
        a = StaticData_Apr20()
        seperated = a[12:16]
        print(seperated)
        if (seperated[3] == "USD"):
            pass
        elif (seperated[3] == "EUR"):
            pass
        elif (seperated[3] == "GBP"):
            pass
        else:
            print("Invalid Opreration")
            continue
        print("FX_Euro_Gbp Datas Information:")
        time.sleep(2)
        b = FX_EURO_GBP()
        print(b)
        print("Time Series Datas Information:")
        time.sleep(2)
        TimeSeriesData_Apr15()
        print("Do you want to Calculate?")
        selection = input("Y/N :")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            continue
        else:
            print("Invalid Selection")
    elif (operations == 4):
        pandas_series[operations]
        print("Selected Index: {} , Product : {}".format(operations, pandas_series.index[operations]))
        print("Static Datas Information:")
        time.sleep(2)
        a = StaticData_Apr20()
        seperated = a[16:20]
        print(seperated)
        if (seperated[3] == "USD"):
            pass
        elif (seperated[3] == "EUR"):
            pass
        elif (seperated[3] == "GBP"):
            pass
        else:
            print("Invalid Opreration")
            continue
        print("FX_Euro_Gbp Datas Information:")
        time.sleep(2)
        b = FX_EURO_GBP()
        print(b)
        print("Time Series Datas Information:")
        time.sleep(2)
        TimeSeriesData_Apr15()
        print("Do you want to Calculate?")
        selection = input("Y/N :")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            continue
        else:
            print("Invalid Selection")
    elif (operations ==5):
        pandas_series[operations]
        print("Selected Index: {} , Product : {}".format(operations, pandas_series.index[operations]))
        print("Static Datas Information:")
        time.sleep(2)
        a = StaticData_Apr20()
        seperated = a[20:24]
        print(seperated)
        if (seperated[3] == "USD"):
            pass
        elif (seperated[3] == "EUR"):
            pass
        elif (seperated[3] == "GBP"):
            pass
        else:
            print("Invalid Opreration")
            continue
        print("FX_Euro_Gbp Datas Information:")
        time.sleep(2)
        b = FX_EURO_GBP()
        print(b)
        print("Time Series Datas Information:")
        time.sleep(2)
        TimeSeriesData_Apr15()
        print("Do you want to Calculate?")
        selection = input("Y/N :")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            continue
        else:
            print("Invalid Selection")
    elif (operations == 6):
        pandas_series[operations]
        print("Selected Index: {} , Product : {}".format(operations, pandas_series.index[operations]))
        print("Static Datas Information:")
        time.sleep(2)
        a = StaticData_Apr20()
        seperated = a[24:28]
        print(seperated)
        if (seperated[3] == "USD"):
            pass
        elif (seperated[3] == "EUR"):
            pass
        elif (seperated[3] == "GBP"):
            pass
        else:
            print("Invalid Opreration")
            continue
        print("FX_Euro_Gbp Datas Information:")
        time.sleep(2)
        b = FX_EURO_GBP()
        print(b)

        print("Time Series Datas Information:")
        time.sleep(2)
        TimeSeriesData_Apr15()
        print("Do you want to Calculate?")
        selection = input("Y/N :")
        if (selection == "Y"):
            pass
        elif (selection == "N"):
            continue
        else:
            print("Invalid Selection")
    elif(operations == "Q"):
        break
    else:
        print("Invalid Operations Please Select A Correct Operation For Calculating Stock..")
    

